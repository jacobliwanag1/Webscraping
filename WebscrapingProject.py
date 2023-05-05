from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
from urllib.request import urlopen, Request






url = 'https://www.coingecko.com/'
# Request in case 404 Forbidden error
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url, headers=headers)

webpage = urlopen(req).read()

soup = BeautifulSoup(webpage, 'html.parser')

print(soup.title.text)

table_rows = soup.findAll("tr")
#print(table_rows[2:20])
wb = xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'

ws['A1'] = 'No.'
ws['B1'] = 'Cryptocurrency'
ws['C1'] = 'Price'
ws['D1'] = '24h percent'
ws['E1'] = 'Mkt cap'
ws['F1'] = 'Corresponding price'


for x in range(1,6):
    td = table_rows[x].findAll('td')
    no = td[1].text
    cryptoname = td[2].text.strip('\n').strip('\n')
    price = float(td[3].text.strip('\n').strip('\n').replace(",","").replace("$",""))
    daily = float(td[5].text.replace("%","").strip('\n'))
    marketcap = float(td[8].text.replace(",","").replace("$",""))

    corr_price = round(price*(1+(daily/100)),2)

    ws['A' + str(x+1)] = no +'' 
    ws['B' + str(x+1)] = cryptoname
    ws['C' + str(x+1)] = price
    ws['D' + str(x+1)] = str(daily) + '%'
    ws['E' + str(x+1)] = marketcap
    ws['F' + str(x+1)] = str(corr_price)
    
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 25

header_font = Font(size=16, bold=True)

for cell in ws[1:1]:
    cell.font = header_font

wb.save("Cryptocurrencies.xlsx")


quote = "The value of Etherium and Bitcoin are within the $5 range of its current value"

import keys
from twilio.rest import Client


client = Client(keys.accountSID, keys.authToken)

TwilioNumber = ''

mycellphone = ''

if ws['F2']-ws['C2']<5:
    textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber,
                                     body=quote)
   


